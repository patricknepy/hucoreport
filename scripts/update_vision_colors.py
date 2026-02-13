"""
Script pour mettre à jour les colonnes Vision Client et Vision Interne
avec le nouveau système de niveaux et couleurs.

Mapping:
- WARNING -> CRITIQUE (violet #7030A0, texte blanc)
- A AMELIORER -> WARNING (orange #FFC000, texte noir)
- BON -> BON (vert #92D050, texte noir)
- NON DEFINI/vide -> NON DEFINI (gris #D9D9D9, texte noir)
- EXCELLENT -> EXCELLENT (bleu #00B0F0, texte noir)
"""

import sys
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment

# Couleurs
COLORS = {
    'CRITIQUE': {'fill': 'FF7030A0', 'font': 'FFFFFFFF'},  # Violet, texte blanc
    'WARNING': {'fill': 'FFFFC000', 'font': 'FF000000'},   # Orange, texte noir
    'BON': {'fill': 'FF92D050', 'font': 'FF000000'},       # Vert, texte noir
    'EXCELLENT': {'fill': 'FF00B0F0', 'font': 'FF000000'}, # Bleu, texte noir
    'NON DEFINI': {'fill': 'FFD9D9D9', 'font': 'FF000000'}, # Gris, texte noir
}

# Mapping ancien -> nouveau
VALUE_MAPPING = {
    'WARNING': 'CRITIQUE',
    'A AMELIORER': 'WARNING',
    'À AMÉLIORER': 'WARNING',
    'A AMÉLIORER': 'WARNING',
    'À AMELIORER': 'WARNING',
    'BON': 'BON',
    'EXCELLENT': 'EXCELLENT',
    'NON DEFINI': 'NON DEFINI',
    'NON DÉFINI': 'NON DEFINI',
    '': 'NON DEFINI',
    None: 'NON DEFINI',
}


def find_column_by_header(sheet, header_row, search_terms):
    """Trouve une colonne par son en-tête (insensible à la casse)."""
    for col in range(1, sheet.max_column + 1):
        cell_value = sheet.cell(row=header_row, column=col).value
        if cell_value:
            cell_str = str(cell_value).upper().strip()
            for term in search_terms:
                if term.upper() in cell_str:
                    return col
    return None


def apply_style(cell, level):
    """Applique le style (couleur de fond + texte) selon le niveau."""
    if level not in COLORS:
        level = 'NON DEFINI'

    color_info = COLORS[level]
    cell.fill = PatternFill(start_color=color_info['fill'],
                            end_color=color_info['fill'],
                            fill_type='solid')
    cell.font = Font(color=color_info['font'], bold=True)
    cell.alignment = Alignment(horizontal='center', vertical='center')


def process_sheet(sheet, header_row=3):
    """Traite un onglet du fichier Excel."""
    # Trouver les colonnes Vision Client et Vision Interne
    vision_client_col = find_column_by_header(sheet, header_row, ['VISION CLIENT'])
    vision_interne_col = find_column_by_header(sheet, header_row, ['VISION INTERNE', 'VISION METIER'])

    changes = {'client': 0, 'interne': 0}

    if vision_client_col:
        print(f"  Vision Client trouvée colonne {vision_client_col}")
        for row in range(header_row + 1, sheet.max_row + 1):
            cell = sheet.cell(row=row, column=vision_client_col)
            old_value = str(cell.value).strip().upper() if cell.value else ''

            # Mapper vers nouvelle valeur
            new_value = VALUE_MAPPING.get(old_value, 'NON DEFINI')

            # Appliquer nouvelle valeur et style
            cell.value = new_value
            apply_style(cell, new_value)
            changes['client'] += 1
    else:
        print("  Vision Client non trouvée")

    if vision_interne_col:
        print(f"  Vision Interne trouvée colonne {vision_interne_col}")
        for row in range(header_row + 1, sheet.max_row + 1):
            cell = sheet.cell(row=row, column=vision_interne_col)
            old_value = str(cell.value).strip().upper() if cell.value else ''

            # Mapper vers nouvelle valeur
            new_value = VALUE_MAPPING.get(old_value, 'NON DEFINI')

            # Appliquer nouvelle valeur et style
            cell.value = new_value
            apply_style(cell, new_value)
            changes['interne'] += 1
    else:
        print("  Vision Interne non trouvée")

    return changes


def main(input_file, output_file=None):
    """Fonction principale."""
    input_path = Path(input_file)

    if not input_path.exists():
        print(f"Erreur: Le fichier {input_file} n'existe pas")
        return False

    if output_file is None:
        # Créer un nom de fichier de sortie
        output_file = input_path.parent / f"{input_path.stem}_updated{input_path.suffix}"

    print(f"Chargement de {input_file}...")
    wb = load_workbook(input_file)

    total_changes = {'client': 0, 'interne': 0}

    for sheet_name in wb.sheetnames:
        # Ignorer les onglets qui ne sont pas des semaines (S01, S02, etc.)
        if not (sheet_name.startswith('S') and any(c.isdigit() for c in sheet_name)):
            print(f"Onglet '{sheet_name}' ignoré (pas une semaine)")
            continue

        print(f"\nTraitement de l'onglet '{sheet_name}'...")
        sheet = wb[sheet_name]
        changes = process_sheet(sheet)
        total_changes['client'] += changes['client']
        total_changes['interne'] += changes['interne']

    print(f"\nSauvegarde vers {output_file}...")
    wb.save(output_file)

    print(f"\nTerminé!")
    print(f"  - Cellules Vision Client modifiées: {total_changes['client']}")
    print(f"  - Cellules Vision Interne modifiées: {total_changes['interne']}")

    return True


if __name__ == '__main__':
    if len(sys.argv) < 2:
        print("Usage: python update_vision_colors.py <fichier_excel> [fichier_sortie]")
        sys.exit(1)

    input_file = sys.argv[1]
    output_file = sys.argv[2] if len(sys.argv) > 2 else None

    success = main(input_file, output_file)
    sys.exit(0 if success else 1)
