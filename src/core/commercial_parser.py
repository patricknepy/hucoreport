"""
Parser pour le fichier input.xlsx (données commerciales).

Ce module lit les données du pipeline commercial et calcule
les indicateurs de vente (Taux Régie/Build, Pipeline pondéré).
"""

import openpyxl
from pathlib import Path
from typing import Dict, List, Any, Optional
from datetime import datetime
import re
import logging

from .paths import get_application_path

logger = logging.getLogger(__name__)


class CommercialParser:
    """Parser de fichier Excel Input Commercial (input.xlsx)."""

    # Pondérations du pipeline par statut
    PIPELINE_WEIGHTS = {
        'signed': 1.0,      # 100% - Contrat signé
        'agreed': 0.8,      # 80% - Accord verbal
        'a': 0.8,           # Alias pour Agreed
        'likely': 0.5,      # 50% - Probable
        'l': 0.5,           # Alias pour Likely
        'specul': 0.2,      # 20% - Spéculatif
        'spec': 0.2,        # Alias pour Specul
    }

    # Types d'offres pour catégorisation
    REGIE_KEYWORDS = ['régie', 'regie', 'tma', 'maintenance']
    BUILD_KEYWORDS = ['dev', 'build', 'forfait', 'projet', 'consulting']

    def __init__(self):
        self.workbook = None
        self.file_path = None
        self.data = []

    def load_file(self, excel_path: str) -> bool:
        """
        Charge le fichier Excel commercial.

        Args:
            excel_path: Chemin vers le fichier Excel (input.xlsx)

        Returns:
            True si chargement réussi
        """
        try:
            self.file_path = Path(excel_path)
            self.workbook = openpyxl.load_workbook(excel_path, data_only=True)
            logger.info(f"Fichier commercial chargé : {self.file_path.name}")
            return True
        except Exception as e:
            logger.error(f"Erreur chargement fichier commercial: {e}")
            return False

    def _detect_header_row(self, ws) -> int:
        """
        Détecte la ligne d'en-tête en cherchant 'Client'.

        Returns:
            Numéro de ligne de l'en-tête (1-indexed)
        """
        for row in range(1, 20):  # Chercher dans les 20 premières lignes
            for col in range(1, 10):
                cell = ws.cell(row=row, column=col).value
                if cell and 'client' in str(cell).lower():
                    logger.info(f"En-tête détecté ligne {row}")
                    return row
        return 5  # Fallback par défaut

    def _detect_columns(self, ws, header_row: int) -> Dict[str, int]:
        """
        Détecte les colonnes par leur nom d'en-tête.

        Returns:
            Dict {nom_champ: index_colonne}
        """
        column_mapping = {}

        # Mots-clés pour chaque colonne
        keywords = {
            'client': ['client'],
            'offre': ['offre', 'offer', 'type'],
            'project': ['project', 'projet'],
            'bu': ['bu'],
            'id_project': ['id_project', 'id projet', 'code'],
            'status': ['status', 'statut'],
            'tjm': ['tjm', 'taux'],
            'days': ['days', 'jours', 'nb jours'],
            'rev2025': ['rev2025', 'rev 2025', 'ca 2025'],
        }

        # Scanner l'en-tête
        for col in range(1, 50):
            cell = ws.cell(row=header_row, column=col).value
            if cell:
                header = str(cell).lower().strip()

                for field, kws in keywords.items():
                    if field not in column_mapping:
                        for kw in kws:
                            if kw in header:
                                column_mapping[field] = col
                                logger.debug(f"Colonne {field} -> {col} ({header})")
                                break

        # Détecter les colonnes mensuelles (Jan, Feb, Mar...)
        months = ['jan', 'feb', 'mar', 'apr', 'may', 'jun',
                  'jul', 'aug', 'sep', 'oct', 'nov', 'dec']
        for col in range(1, 50):
            cell = ws.cell(row=header_row, column=col).value
            if cell:
                header = str(cell).lower().strip()
                for month in months:
                    if header == month or header == f'{month}r':
                        key = f'month_{month}' if header == month else f'month_{month}_r'
                        column_mapping[key] = col

        return column_mapping

    def parse_sheet(self, sheet_name: str = "HUMAN'S") -> List[Dict[str, Any]]:
        """
        Parse l'onglet des données commerciales.

        Args:
            sheet_name: Nom de l'onglet à parser

        Returns:
            Liste des lignes commerciales
        """
        if not self.workbook:
            raise ValueError("Aucun fichier chargé")

        # Trouver l'onglet (essayer plusieurs noms)
        possible_names = [sheet_name, "HUMAN'S", "HUMANS", "Data", "Données"]
        ws = None
        for name in possible_names:
            if name in self.workbook.sheetnames:
                ws = self.workbook[name]
                break

        if ws is None:
            # Prendre le premier onglet par défaut
            ws = self.workbook.active
            logger.warning(f"Onglet {sheet_name} non trouvé, utilisation de: {ws.title}")

        # Détecter en-tête et colonnes
        header_row = self._detect_header_row(ws)
        columns = self._detect_columns(ws, header_row)

        logger.info(f"Colonnes détectées: {list(columns.keys())}")

        # Parser les données
        data = []
        for row in range(header_row + 1, ws.max_row + 1):
            # Vérifier si ligne non vide (client renseigné)
            client_col = columns.get('client', 1)
            client = ws.cell(row=row, column=client_col).value

            if not client or (isinstance(client, str) and not client.strip()):
                continue

            record = {
                'row': row,
                'client': str(client).strip() if client else None,
                'offre': None,
                'project': None,
                'bu': None,
                'status': None,
                'tjm': None,
                'days': None,
                'rev2025': None,
            }

            # Lire chaque colonne
            for field, col_idx in columns.items():
                if field.startswith('month_'):
                    continue  # Traiter séparément

                value = ws.cell(row=row, column=col_idx).value
                if field in record:
                    record[field] = self._convert_value(value, field)

            # Lire les mois prévisionnels et réalisés
            monthly_planned = {}
            monthly_actual = {}
            months = ['jan', 'feb', 'mar', 'apr', 'may', 'jun',
                      'jul', 'aug', 'sep', 'oct', 'nov', 'dec']

            for month in months:
                if f'month_{month}' in columns:
                    val = ws.cell(row=row, column=columns[f'month_{month}']).value
                    monthly_planned[month] = self._to_number(val)
                if f'month_{month}_r' in columns:
                    val = ws.cell(row=row, column=columns[f'month_{month}_r']).value
                    monthly_actual[month] = self._to_number(val)

            record['monthly_planned'] = monthly_planned
            record['monthly_actual'] = monthly_actual

            data.append(record)

        self.data = data
        logger.info(f"Parsé {len(data)} lignes commerciales")
        return data

    def _convert_value(self, value: Any, field: str) -> Any:
        """Convertit une valeur selon le champ."""
        if value is None or value == '':
            return None

        if field in ['tjm', 'days', 'rev2025']:
            return self._to_number(value)
        else:
            return str(value).strip()

    def _to_number(self, value: Any) -> Optional[float]:
        """Convertit en nombre."""
        if value is None or value == '':
            return None
        try:
            if isinstance(value, (int, float)):
                return float(value)
            return float(str(value).replace(',', '.').replace(' ', ''))
        except:
            return None

    def get_offer_type(self, offre: str) -> str:
        """
        Catégorise une offre en Régie ou Build.

        Args:
            offre: Type d'offre (texte)

        Returns:
            'REGIE' ou 'BUILD'
        """
        if not offre:
            return 'BUILD'  # Par défaut

        offre_lower = offre.lower()

        for kw in self.REGIE_KEYWORDS:
            if kw in offre_lower:
                return 'REGIE'

        return 'BUILD'

    def get_pipeline_weight(self, status: str) -> float:
        """
        Retourne la pondération pipeline pour un statut.

        Args:
            status: Statut (Signed, Agreed, Likely, Specul)

        Returns:
            Pondération (0.2 à 1.0)
        """
        if not status:
            return 0.0

        status_clean = status.lower().strip()
        return self.PIPELINE_WEIGHTS.get(status_clean, 0.0)

    def calculate_kpis(self) -> Dict[str, Any]:
        """
        Calcule tous les KPIs commerciaux.

        Returns:
            Dictionnaire avec tous les KPIs
        """
        if not self.data:
            return {
                'taux_regie': 0,
                'taux_build': 0,
                'pipeline_total': 0,
                'pipeline_pondere': 0,
                'ca_prevu_annuel': 0,
                'ca_realise_ytd': 0,
                'taux_realisation': 0,
                'nb_projets_signed': 0,
                'nb_projets_agreed': 0,
                'nb_projets_likely': 0,
                'nb_projets_specul': 0,
                'tjm_moyen': 0,
            }

        # Compteurs
        count_regie = 0
        count_build = 0
        total_projects = 0

        pipeline_total = 0
        pipeline_pondere = 0

        ca_prevu = 0
        ca_realise = 0

        nb_signed = 0
        nb_agreed = 0
        nb_likely = 0
        nb_specul = 0

        tjm_sum = 0
        tjm_count = 0

        for record in self.data:
            total_projects += 1

            # Taux Régie/Build
            offer_type = self.get_offer_type(record.get('offre'))
            if offer_type == 'REGIE':
                count_regie += 1
            else:
                count_build += 1

            # Pipeline - utiliser rev2025 comme CA prévu
            rev = record.get('rev2025') or 0
            status = record.get('status', '')
            weight = self.get_pipeline_weight(status)

            pipeline_total += rev
            pipeline_pondere += rev * weight

            # CA prévu = rev2025 (prévisionnel annuel)
            ca_prevu += rev

            # Comptage par statut
            if status:
                status_lower = status.lower().strip()
                if status_lower == 'signed':
                    nb_signed += 1
                elif status_lower in ['agreed', 'a']:
                    nb_agreed += 1
                elif status_lower in ['likely', 'l']:
                    nb_likely += 1
                elif status_lower in ['specul', 'spec']:
                    nb_specul += 1

            # CA réalisé = somme des mois réalisés (colonnes JanR, FebR, etc.)
            monthly_actual = record.get('monthly_actual', {})
            for month, val in monthly_actual.items():
                if val and val > 0:
                    ca_realise += val

            # TJM moyen
            tjm = record.get('tjm')
            if tjm and tjm > 0:
                tjm_sum += tjm
                tjm_count += 1

        # Calculer les taux
        taux_regie = round((count_regie / total_projects * 100) if total_projects > 0 else 0, 1)
        taux_build = round((count_build / total_projects * 100) if total_projects > 0 else 0, 1)

        # Taux de réalisation : CA réalisé / CA prévu (plafonné à 200% pour éviter les aberrations)
        if ca_prevu > 0 and ca_realise > 0:
            taux_realisation = round((ca_realise / ca_prevu * 100), 1)
            # Plafonner à 200% si aberrant
            if taux_realisation > 200:
                taux_realisation = 0  # Données incohérentes, afficher 0
        else:
            taux_realisation = 0

        tjm_moyen = round(tjm_sum / tjm_count) if tjm_count > 0 else 0

        logger.info(f"KPIs commerciaux: pipeline={pipeline_total}, pondéré={pipeline_pondere}, "
                   f"prévu={ca_prevu}, réalisé={ca_realise}, taux={taux_realisation}%")

        return {
            'taux_regie': taux_regie,
            'taux_build': taux_build,
            'pipeline_total': round(pipeline_total, 0),
            'pipeline_pondere': round(pipeline_pondere, 0),
            'ca_prevu_annuel': round(ca_prevu, 0),
            'ca_realise_ytd': round(ca_realise, 0),
            'taux_realisation': taux_realisation,
            'nb_projets_signed': nb_signed,
            'nb_projets_agreed': nb_agreed,
            'nb_projets_likely': nb_likely,
            'nb_projets_specul': nb_specul,
            'tjm_moyen': tjm_moyen,
            'total_projets': total_projects,
        }

    def close(self):
        """Ferme le fichier Excel."""
        if self.workbook:
            self.workbook.close()
            self.workbook = None
