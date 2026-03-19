"""
Lecture et parsing du fichier Excel.

Ce module lit les données du fichier Excel, les valide et prépare
la simulation d'import.
"""

import openpyxl
from pathlib import Path
from typing import Dict, List, Any, Optional, Tuple
from datetime import datetime, date, timedelta
import re
import logging
import json
import sys

from .paths import get_application_path

logger = logging.getLogger(__name__)


class ExcelParser:
    """Parser de fichier Excel Suivi Hebdo Projet."""

    def __init__(self, schema_path: str = "config/excel_schema.json"):
        """
        Initialise le parser avec le schéma.

        Args:
            schema_path: Chemin vers le fichier JSON du schéma
        """
        # Utiliser un chemin absolu basé sur l'emplacement de l'exe
        app_path = get_application_path()
        self.schema_path = app_path / schema_path
        self.schema = self._load_schema()
        self.workbook = None
        self.file_path = None

    def _load_schema(self) -> Dict:
        """Charge le schéma JSON."""
        with open(self.schema_path, 'r', encoding='utf-8') as f:
            return json.load(f)

    def load_file(self, excel_path: str):
        """
        Charge le fichier Excel.

        Args:
            excel_path: Chemin vers le fichier Excel
        """
        self.file_path = Path(excel_path)
        self.workbook = openpyxl.load_workbook(excel_path, data_only=True)
        logger.info(f"Fichier charge : {self.file_path.name}")

    def get_available_weeks(self) -> List[int]:
        """
        Retourne la liste des semaines disponibles dans le fichier.

        Returns:
            Liste de numéros de semaines (ex: [45, 46, 47, 48])
        """
        if not self.workbook:
            return []

        sheet_pattern = re.compile(self.schema['sheet_pattern'])
        weeks = []

        for sheet_name in self.workbook.sheetnames:
            if sheet_pattern.match(sheet_name):
                week_num = int(sheet_name[1:])  # S48 → 48
                weeks.append(week_num)

        return sorted(weeks)

    def _convert_to_boolean(self, value: Any) -> bool:
        """
        Convertit une valeur Excel en booléen.

        Args:
            value: Valeur Excel (X, x, vide, etc.)

        Returns:
            True si X/x, False sinon
        """
        if value is None or value == '':
            return False

        if isinstance(value, str):
            return value.strip().upper() == 'X'

        return bool(value)

    def _convert_to_date(self, value: Any) -> Optional[str]:
        """
        Convertit une valeur Excel en date ISO (YYYY-MM-DD).

        Gère plusieurs formats :
        - Date complète : JJ/MM/AAAA, AAAA-MM-JJ
        - Mois seul : 2025-11, Novembre 2025 → dernier jour du mois
        - Semaine : S48, semaine 48 → dernier jour (dimanche)

        Args:
            value: Valeur Excel (datetime, str, etc.)

        Returns:
            Date au format ISO ou None
        """
        if value is None or value == '':
            return None

        # Si déjà un objet datetime
        if isinstance(value, datetime):
            return value.date().isoformat()

        if isinstance(value, date):
            return value.isoformat()

        # Si string, essayer de parser
        if isinstance(value, str):
            value_clean = value.strip()

            # 1. Formats de date complète
            for fmt in ['%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y', '%d/%m/%y']:
                try:
                    dt = datetime.strptime(value_clean, fmt)
                    return dt.date().isoformat()
                except ValueError:
                    continue

            # 2. Format MOIS seul : "2025-11" ou "11/2025"
            # → Retourner le dernier jour du mois
            for fmt in ['%Y-%m', '%m/%Y']:
                try:
                    dt = datetime.strptime(value_clean, fmt)
                    # Calculer le dernier jour du mois
                    if dt.month == 12:
                        last_day = date(dt.year, 12, 31)
                    else:
                        last_day = date(dt.year, dt.month + 1, 1) - timedelta(days=1)
                    return last_day.isoformat()
                except ValueError:
                    continue

            # 3. Format SEMAINE : "S48", "semaine 48", "sem 48"
            # → Retourner le dimanche de cette semaine
            week_match = re.match(r'^[Ss]?(?:emaine\s*)?(\d{1,2})$', value_clean)
            if week_match:
                week_num = int(week_match.group(1))
                # Calculer le dimanche de cette semaine
                # (On suppose l'année en cours)
                current_year = datetime.now().year

                # Premier jour de l'année
                jan_1 = date(current_year, 1, 1)

                # Trouver le premier lundi de l'année
                days_to_monday = (7 - jan_1.weekday()) % 7
                if days_to_monday == 0 and jan_1.weekday() != 0:
                    days_to_monday = 7
                first_monday = jan_1 + timedelta(days=days_to_monday)

                # Calculer le dimanche de la semaine N
                target_week_start = first_monday + timedelta(weeks=week_num - 1)
                sunday = target_week_start + timedelta(days=6)

                return sunday.isoformat()

            # 4. Mois en texte : "Novembre 2025", "Nov 2025"
            mois_fr = {
                'janvier': 1, 'jan': 1,
                'février': 2, 'fev': 2, 'fevrier': 2,
                'mars': 3, 'mar': 3,
                'avril': 4, 'avr': 4,
                'mai': 5,
                'juin': 6,
                'juillet': 7, 'juil': 7,
                'août': 8, 'aout': 8,
                'septembre': 9, 'sep': 9, 'sept': 9,
                'octobre': 10, 'oct': 10,
                'novembre': 11, 'nov': 11,
                'décembre': 12, 'dec': 12, 'decembre': 12
            }

            for mois_nom, mois_num in mois_fr.items():
                pattern = rf'{mois_nom}\s*(\d{{4}})'
                match = re.search(pattern, value_clean.lower())
                if match:
                    year = int(match.group(1))
                    # Dernier jour du mois
                    if mois_num == 12:
                        last_day = date(year, 12, 31)
                    else:
                        last_day = date(year, mois_num + 1, 1) - timedelta(days=1)
                    return last_day.isoformat()

        logger.warning(f"Impossible de convertir en date : {value}")
        return None

    def _convert_to_integer(self, value: Any) -> Optional[int]:
        """
        Convertit une valeur Excel en entier.

        Args:
            value: Valeur Excel

        Returns:
            Entier ou None
        """
        if value is None or value == '':
            return None

        try:
            # Si c'est un float, arrondir
            if isinstance(value, float):
                return int(round(value))

            # Si c'est un str, vérifier d'abord les erreurs Excel
            if isinstance(value, str):
                value_stripped = value.strip()
                if value_stripped == '':
                    return None
                # Détecter les erreurs Excel (#VALUE!, #REF!, #N/A, etc.)
                if value_stripped.startswith('#') and ('!' in value_stripped or value_stripped == '#N/A'):
                    logger.warning(f"Erreur Excel detectee : {value_stripped}. La cellule contient une formule invalide. Veuillez corriger le fichier Excel.")
                    return None
                return int(float(value_stripped))

            return int(value)
        except (ValueError, TypeError):
            logger.warning(f"Impossible de convertir en entier : {value}")
            return None

    def _convert_to_float(self, value: Any) -> Optional[float]:
        """
        Convertit une valeur Excel en float.

        Args:
            value: Valeur Excel

        Returns:
            Float ou None
        """
        if value is None or value == '':
            return None

        try:
            # Si c'est déjà un nombre
            if isinstance(value, (int, float)):
                return float(value)

            # Si c'est un str
            if isinstance(value, str):
                value_stripped = value.strip()
                if value_stripped == '':
                    return None
                # Détecter les erreurs Excel
                if value_stripped.startswith('#') and ('!' in value_stripped or value_stripped == '#N/A'):
                    logger.warning(f"Erreur Excel detectee : {value_stripped}.")
                    return None
                # Gérer la virgule française comme séparateur décimal
                value_normalized = value_stripped.replace(',', '.')
                return float(value_normalized)

            return float(value)
        except (ValueError, TypeError):
            logger.warning(f"Impossible de convertir en float : {value}")
            return None

    def _convert_to_text(self, value: Any) -> Optional[str]:
        """
        Convertit une valeur Excel en texte.

        Args:
            value: Valeur Excel

        Returns:
            Texte ou None
        """
        if value is None or value == '':
            return None

        return str(value).strip()

    def _normalize_status(self, value: Any) -> Optional[str]:
        """
        Normalise les statuts de projet.

        Mapping:
        - "Actif" → "EN COURS"
        - "Pause" → "PAUSE"
        - "Non Actif" → "TERMINÉ"
        - "À Venir" → "À VENIR"

        Args:
            value: Valeur Excel du statut

        Returns:
            Statut normalisé
        """
        if value is None or value == '':
            return None

        status_mapping = {
            'actif': 'EN COURS',
            'pause': 'PAUSE',
            'non actif': 'TERMINÉ',
            'à venir': 'À VENIR',
            'terminé': 'TERMINÉ',
            'en cours': 'EN COURS'
        }

        normalized = str(value).strip().lower()
        return status_mapping.get(normalized, str(value).strip())

    def _convert_value(self, value: Any, expected_type: str) -> Any:
        """
        Convertit une valeur selon le type attendu.

        Args:
            value: Valeur Excel
            expected_type: Type attendu (INTEGER, DATE, TEXT, BOOLEAN, FLOAT)

        Returns:
            Valeur convertie
        """
        if expected_type == 'INTEGER':
            return self._convert_to_integer(value)
        elif expected_type == 'FLOAT':
            return self._convert_to_float(value)
        elif expected_type == 'DATE':
            return self._convert_to_date(value)
        elif expected_type == 'BOOLEAN':
            return self._convert_to_boolean(value)
        elif expected_type == 'TEXT':
            return self._convert_to_text(value)
        else:
            return value

    def _detect_columns_from_header(self, ws, header_row: int) -> Dict[str, Dict]:
        """
        Détecte dynamiquement les colonnes en lisant la ligne d'en-tête.

        INTELLIGENCE DE DÉTECTION :
        1. Correspondance exacte avec les noms du schéma
        2. Correspondance par mots-clés pour les colonnes critiques
        3. Fallback sur correspondance partielle

        Args:
            ws: Worksheet openpyxl
            header_row: Numéro de ligne de l'en-tête

        Returns:
            Dict {col_index: {'db_field': ..., 'type': ...}}
        """
        detected_mapping = {}

        # MOTS-CLÉS CRITIQUES pour détecter les colonnes importantes
        # Format: 'db_field': [(mots_requis), type, (mots_exclus)]
        critical_keywords = {
            'id_projet': (['number'], 'INTEGER', []),
            'status': (['statut'], 'TEXT', []),
            'bu': (['bu'], 'TEXT', ['club']),
            'client_name': (['client'], 'TEXT', ['relation', 'satisfaction', 'échange', 'exchange']),
            'project_manager': (['chef', 'projet'], 'TEXT', ['direction']),
            'vision_client': (['vision', 'client'], 'TEXT', ['commentaire']),
            'vision_internal': (['vision', 'interne'], 'TEXT', ['commentaire']),
            'comment_vision_client': (['commentaire', 'vision', 'client'], 'TEXT', []),
            'comment_vision_internal': (['commentaire', 'vision', 'interne'], 'TEXT', []),
            'upsell': (['upsell'], 'TEXT', []),
            'crosssell': (['cros', 'sell'], 'TEXT', []),
            'next_actor': (['acteur'], 'TEXT', []),
            'dlic': (['dlic'], 'DATE', []),
            'dli': (['dli', 'limite', 'interne'], 'DATE', ['dlic']),
            'action_description': (['action'], 'TEXT', ['satisfaction']),
            'technical_lead': (['dév', 'principal'], 'TEXT', []),
            'project_director': (['direction', 'projet'], 'TEXT', []),
            'project_phase': (['phase'], 'TEXT', []),
            'contract_type': (['type', 'contrat'], 'TEXT', []),
            'project_code': (['code', 'projet'], 'TEXT', []),
            'description': (['description'], 'TEXT', []),
            'last_client_exchange': (['dernier', 'échange', 'client'], 'DATE', []),
            'next_client_exchange': (['prochain', 'échange', 'client'], 'DATE', []),
            'news_project': (['actualité', 'projet'], 'TEXT', []),
            'news_commercial': (['actualité', 'commerciale'], 'TEXT', []),
            'news_technical': (['actualité', 'technique'], 'TEXT', []),
            'risk_identified': (['risque', 'identifié'], 'TEXT', []),
            'nps_commercial': (['relation', 'client', 'commer'], 'INTEGER', []),
            'nps_project': (['relation', 'client', 'projet'], 'INTEGER', []),
            'next_milestone_date': (['date', 'jalon'], 'DATE', []),
            'next_milestone_object': (['objet', 'jalon'], 'TEXT', []),
            'days_sold': (['jours', 'vendus'], 'INTEGER', []),
            'days_dispositif_monthly': (['dispositif', 'mensuel'], 'INTEGER', []),
            'dispositif_expandable': (['dispositif', 'augmentable'], 'TEXT', []),
            'days_forfait': (['jours', 'forfait'], 'INTEGER', []),
            'days_to_sign': (['jours', 'signer'], 'INTEGER', []),
            'days_facturable_main': (['temps', 'facturable', 'main'], 'FLOAT', []),
            'start_date': (['date', 'démarrage'], 'DATE', []),
            'data_remarks': (['remarques', 'data'], 'TEXT', []),
            'remarks_3months': (['objectifs', '3', 'mois'], 'TEXT', []),
            'remarks_6months': (['objectifs', '6', 'mois'], 'TEXT', []),
            'remarks_1year': (['objectifs', '1', 'an'], 'TEXT', []),
            'potential_new_projects': (['projet'], 'BOOLEAN', ['code', 'chef', 'direction', 'phase', 'actualité', 'risque', 'relation']),
            'potential_maintenance': (['maint'], 'BOOLEAN', []),
            'potential_hosting': (['héberg', 'herbeg'], 'BOOLEAN', []),
            'potential_infra': (['infra'], 'BOOLEAN', []),
            'potential_consulting': (['conseil'], 'BOOLEAN', []),
        }

        # Scanner la ligne d'en-tête (colonnes A à BZ)
        all_headers = {}
        for col_idx in range(1, 79):
            cell_value = ws.cell(row=header_row, column=col_idx).value
            if cell_value:
                header_clean = str(cell_value).strip().lower().replace('\n', ' ').replace('  ', ' ')
                all_headers[col_idx] = header_clean

        # Détecter chaque colonne critique par mots-clés
        for db_field, (required_words, field_type, excluded_words) in critical_keywords.items():
            best_match = None
            best_score = 0

            for col_idx, header in all_headers.items():
                # Vérifier exclusions d'abord
                excluded = False
                for excl in excluded_words:
                    if excl in header:
                        excluded = True
                        break
                if excluded:
                    continue

                # Compter mots requis présents
                score = sum(1 for word in required_words if word in header)

                # Match si tous les mots requis sont présents
                if score == len(required_words) and score > best_score:
                    # Vérifier pas déjà assigné
                    if col_idx not in detected_mapping:
                        best_match = col_idx
                        best_score = score

            if best_match:
                detected_mapping[best_match] = {
                    'db_field': db_field,
                    'type': field_type,
                    'header_found': ws.cell(row=header_row, column=best_match).value
                }
                col_letter = openpyxl.utils.get_column_letter(best_match)
                logger.info(f"  {db_field} -> colonne {col_letter} ({all_headers[best_match][:30]})")

        # Log des colonnes critiques non trouvées
        critical_fields = ['id_projet', 'status', 'bu', 'client_name', 'project_manager',
                          'vision_client', 'vision_internal']
        for field in critical_fields:
            if not any(m['db_field'] == field for m in detected_mapping.values()):
                logger.warning(f"  ATTENTION: {field} non trouvé dans cet onglet!")

        return detected_mapping

    def parse_sheet(self, week_number: int) -> List[Dict[str, Any]]:
        """
        Parse un onglet de semaine et extrait les projets.

        DÉTECTION DYNAMIQUE : Les colonnes sont détectées par leur nom d'en-tête,
        pas par leur position fixe. Cela permet de gérer les fichiers où les
        colonnes ont été réorganisées.

        Args:
            week_number: Numéro de semaine (ex: 48)

        Returns:
            Liste de dictionnaires représentant les projets
        """
        if not self.workbook:
            raise ValueError("Aucun fichier chargé. Appelez load_file() d'abord.")

        # Essayer plusieurs formats : S2, S02, S002...
        possible_names = [
            f"S{week_number}",        # S2, S48
            f"S{week_number:02d}",    # S02, S48
            f"S{week_number:03d}"     # S002, S048
        ]

        sheet_name = None
        for name in possible_names:
            if name in self.workbook.sheetnames:
                sheet_name = name
                break

        if sheet_name is None:
            raise ValueError(f"Onglet S{week_number} introuvable (essayé: {', '.join(possible_names)})")

        ws = self.workbook[sheet_name]
        projects = []

        header_row = self.schema['header_row']
        data_start = self.schema['data_start_row']

        # DÉTECTION DYNAMIQUE de la dernière ligne avec données
        # On ne se limite plus à data_end_row du schéma (54) car les fichiers peuvent avoir plus de lignes
        data_end = ws.max_row
        logger.info(f"Onglet {sheet_name}: lecture lignes {data_start} à {data_end} (max_row détecté)")

        # DÉTECTION DYNAMIQUE des colonnes par leur en-tête
        logger.info(f"=== Détection colonnes pour {sheet_name} ===")
        dynamic_mapping = self._detect_columns_from_header(ws, header_row)

        if not dynamic_mapping:
            # Fallback sur le mapping fixe si la détection échoue
            logger.warning(f"Détection dynamique échouée pour {sheet_name}, utilisation du mapping fixe")
            for col_letter, col_schema in self.schema['columns'].items():
                if col_schema.get('db_field') and not col_schema.get('ignored', False):
                    col_idx = openpyxl.utils.column_index_from_string(col_letter)
                    dynamic_mapping[col_idx] = {
                        'db_field': col_schema['db_field'],
                        'type': col_schema.get('type', 'TEXT')
                    }
        else:
            # Résumé des colonnes critiques détectées
            critical = ['vision_client', 'vision_internal', 'project_manager', 'status', 'next_actor', 'dlic']
            found_critical = [f for f in critical if any(m['db_field'] == f for m in dynamic_mapping.values())]
            logger.info(f"Onglet {sheet_name} : {len(dynamic_mapping)} colonnes, critiques: {', '.join(found_critical)}")

        # Lire les lignes de données
        # On s'arrête à la première ligne complètement vide (fin du tableau)
        consecutive_empty = 0
        for row_idx in range(data_start, data_end + 1):
            # Vérifier si ligne vide AVANT de parser
            bu_cell = None
            client_cell = None
            for col_idx, col_info in dynamic_mapping.items():
                if col_info['db_field'] == 'bu':
                    bu_cell = ws.cell(row=row_idx, column=col_idx).value
                elif col_info['db_field'] == 'client_name':
                    client_cell = ws.cell(row=row_idx, column=col_idx).value

            # Si BU et Client sont vides, c'est peut-être la fin du tableau
            bu_empty = not bu_cell or (isinstance(bu_cell, str) and not bu_cell.strip())
            client_empty = not client_cell or (isinstance(client_cell, str) and not client_cell.strip())

            if bu_empty and client_empty:
                consecutive_empty += 1
                if consecutive_empty >= 3:  # 3 lignes vides consécutives = fin du tableau
                    logger.info(f"Fin du tableau détectée ligne {row_idx - 2}")
                    break
                continue
            else:
                consecutive_empty = 0

            project = {'week_number': week_number}

            # Lire chaque colonne détectée
            for col_idx, col_info in dynamic_mapping.items():
                cell_value = ws.cell(row=row_idx, column=col_idx).value

                db_field = col_info['db_field']
                expected_type = col_info['type']

                # Normaliser statut
                if db_field == 'status':
                    converted_value = self._normalize_status(cell_value)
                else:
                    # Convertir la valeur
                    converted_value = self._convert_value(cell_value, expected_type)

                project[db_field] = converted_value

            # Vérifier APRÈS conversion que les champs obligatoires sont valides
            id_projet = project.get('id_projet')
            bu = project.get('bu')
            client_name = project.get('client_name')

            # RÈGLE STRICTE : BU ET client_name doivent être renseignés
            if not bu or (isinstance(bu, str) and not bu.strip()):
                continue
            if not client_name or (isinstance(client_name, str) and not client_name.strip()):
                continue

            # TOUJOURS utiliser row_idx comme ID unique (la colonne Number peut avoir des doublons)
            # car certains onglets ont plusieurs sections avec numérotation qui redémarre à 1
            project['id_projet'] = row_idx + (week_number * 1000)

            projects.append(project)

        logger.info(f"Onglet {sheet_name} : {len(projects)} projets parsés")
        return projects

    def parse_all_weeks(self) -> Dict[int, List[Dict[str, Any]]]:
        """
        Parse toutes les semaines disponibles.

        Returns:
            Dict {week_number: [projects]}
        """
        all_data = {}
        weeks = self.get_available_weeks()

        for week in weeks:
            projects = self.parse_sheet(week)
            all_data[week] = projects

        total_projects = sum(len(projects) for projects in all_data.values())
        logger.info(f"Total : {total_projects} projets sur {len(weeks)} semaines")

        return all_data

    def simulate_import(self, excel_path: str) -> Dict[str, Any]:
        """
        Simule l'import complet du fichier Excel.

        Cette fonction :
        1. Charge le fichier
        2. Parse toutes les semaines
        3. Calcule des statistiques
        4. Valide les données

        Args:
            excel_path: Chemin vers le fichier Excel

        Returns:
            Dictionnaire avec résultats de simulation
        """
        result = {
            "valid": True,
            "file_name": Path(excel_path).name,
            "file_path": str(Path(excel_path)),
            "weeks_detected": [],
            "total_projects": 0,
            "active_projects": 0,
            "projects_by_week": {},
            "errors": [],
            "warnings": []
        }

        try:
            # 1. Charger le fichier
            self.load_file(excel_path)

            # 2. Détecter les semaines
            weeks = self.get_available_weeks()
            result['weeks_detected'] = weeks

            # Calculer la dernière semaine avec le tri circulaire (6 mois glissants)
            if weeks:
                from datetime import datetime
                current_week = datetime.now().isocalendar()[1]

                def week_distance(week: int) -> int:
                    if week <= current_week:
                        return current_week - week
                    else:
                        return current_week + (52 - week)

                # La plus récente = celle avec la plus petite distance
                result['latest_week'] = min(weeks, key=week_distance)
            else:
                result['latest_week'] = None

            if not weeks:
                result['valid'] = False
                result['errors'].append("Aucune semaine détectée dans le fichier")
                return result

            # 3. Parser toutes les semaines
            all_data = self.parse_all_weeks()

            # 4. Calculer statistiques
            for week, projects in all_data.items():
                result['projects_by_week'][week] = len(projects)
                result['total_projects'] += len(projects)

                # Compter projets actifs (dernière semaine uniquement)
                if week == result['latest_week']:
                    active_count = sum(1 for p in projects if p.get('status') == 'EN COURS')
                    result['active_projects'] = active_count

            # 5. Calculer indicateurs de validation (dernière semaine uniquement)
            latest_projects = all_data[result['latest_week']]

            # Compter DLIC cette semaine et dépassées
            from datetime import datetime, timedelta
            today = datetime.now().date()
            week_start = today - timedelta(days=today.weekday())  # Lundi
            week_end = week_start + timedelta(days=6)  # Dimanche

            dlic_this_week = 0
            dlic_overdue = 0
            actifs_with_actor = 0
            total_actifs = 0

            for project in latest_projects:
                status = project.get('status')
                dlic = project.get('dlic')
                next_actor = project.get('next_actor')
                vision_client = str(project.get('vision_client', '')).upper()
                vision_internal = str(project.get('vision_internal', '')).upper()

                # Compter projets actifs
                if status == 'EN COURS':
                    total_actifs += 1

                    # Acteur identifié (sur TOUS les actifs)
                    if next_actor and next_actor.strip():
                        actifs_with_actor += 1

                    # DLIC cette semaine
                    if dlic:
                        try:
                            from datetime import date
                            if isinstance(dlic, str):
                                dlic_date = date.fromisoformat(dlic)
                            else:
                                dlic_date = dlic

                            # DLIC dans la semaine
                            if week_start <= dlic_date <= week_end:
                                dlic_this_week += 1

                            # DLIC dépassée
                            if dlic_date < today:
                                dlic_overdue += 1
                        except:
                            pass

            # Calculer pourcentage (actifs avec acteur / total actifs)
            pct_with_actor = round((actifs_with_actor / total_actifs * 100) if total_actifs > 0 else 0, 1)

            # Stocker dans result
            result['validation_indicators'] = {
                'dlic_this_week': dlic_this_week,
                'dlic_overdue': dlic_overdue,
                'actifs_with_actor': actifs_with_actor,
                'total_actifs': total_actifs,
                'pct_with_actor': pct_with_actor
            }

            # 6. Stocker les données pour import ultérieur
            self.parsed_data = all_data

            logger.info(f"Simulation terminée : {result['total_projects']} projets, {len(weeks)} semaines")

        except Exception as e:
            result['valid'] = False
            result['errors'].append(f"Erreur lors de la simulation : {str(e)}")
            logger.exception(e)

        return result

    def get_parsed_data(self) -> Optional[Dict[int, List[Dict[str, Any]]]]:
        """
        Retourne les données parsées (après simulate_import).

        Returns:
            Dict {week_number: [projects]} ou None
        """
        return getattr(self, 'parsed_data', None)

    def close(self):
        """Ferme le fichier Excel."""
        if self.workbook:
            self.workbook.close()
            self.workbook = None
            logger.info("Fichier Excel fermé")

