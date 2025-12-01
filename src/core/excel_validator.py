"""
Validation de la structure du fichier Excel.

Ce module vérifie que le fichier Excel correspond exactement au schéma de référence
défini dans config/excel_schema.json.
"""

import json
import openpyxl
from pathlib import Path
from typing import Dict, List, Tuple, Any
import re
import logging
import sys

from .paths import get_application_path

logger = logging.getLogger(__name__)


class ExcelValidator:
    """Validateur de structure de fichier Excel."""
    
    def __init__(self, schema_path: str = "config/excel_schema.json"):
        """
        Initialise le validateur avec le schéma de référence.
        
        Args:
            schema_path: Chemin vers le fichier JSON du schéma
        """
        # Utiliser un chemin absolu basé sur l'emplacement de l'exe
        app_path = get_application_path()
        self.schema_path = app_path / schema_path
        self.schema = self._load_schema()
    
    def _load_schema(self) -> Dict:
        """Charge le schéma JSON."""
        try:
            with open(self.schema_path, 'r', encoding='utf-8') as f:
                return json.load(f)
        except FileNotFoundError:
            logger.error(f"❌ Fichier schéma introuvable : {self.schema_path}")
            raise
        except json.JSONDecodeError as e:
            logger.error(f"❌ Erreur parsing JSON : {e}")
            raise
    
    def _normalize_text(self, text: str) -> str:
        """
        Normalise un texte pour comparaison.
        
        - Supprime les retours à la ligne
        - Supprime espaces multiples
        - Minuscules
        - Supprime accents (optionnel)
        
        Args:
            text: Texte à normaliser
            
        Returns:
            Texte normalisé
        """
        if not text:
            return ""
        
        # Remplacer \n par espace
        text = text.replace('\n', ' ')
        
        # Supprimer espaces multiples
        text = re.sub(r'\s+', ' ', text)
        
        # Trim
        text = text.strip()
        
        return text
    
    def _compare_column_name(self, found: Any, expected: str, alternate_names: List[str] = None) -> bool:
        """
        Compare un nom de colonne trouvé avec le nom attendu.
        
        Args:
            found: Nom trouvé dans le fichier Excel (peut être string, int, float, etc.)
            expected: Nom attendu selon le schéma
            alternate_names: Noms alternatifs acceptés
            
        Returns:
            True si correspondance, False sinon
        """
        # Convertir en string si ce n'est pas déjà le cas (gère le cas "0" qui peut être un int)
        if found is None:
            return False
        
        found_str = str(found).strip() if found is not None else ""
        
        if not found_str:
            return False
        
        # Normaliser
        found_normalized = self._normalize_text(found_str)
        expected_normalized = self._normalize_text(expected)
        
        # Comparaison exacte normalisée
        if found_normalized.lower() == expected_normalized.lower():
            return True
        
        # Comparaison avec noms alternatifs
        if alternate_names:
            for alt in alternate_names:
                alt_str = str(alt).strip()  # Convertir en string aussi
                alt_normalized = self._normalize_text(alt_str)
                if found_normalized.lower() == alt_normalized.lower():
                    return True
        
        return False
    
    def validate_structure(self, excel_path: str) -> Dict[str, Any]:
        """
        Valide la structure complète du fichier Excel.
        
        Args:
            excel_path: Chemin vers le fichier Excel
            
        Returns:
            Dictionnaire avec résultats de validation :
            {
                "valid": True/False,
                "errors": [...],
                "warnings": [...],
                "sheets_found": [...],
                "columns_validated": {...}
            }
        """
        result = {
            "valid": True,
            "errors": [],
            "warnings": [],
            "sheets_found": [],
            "columns_validated": {}
        }
        
        try:
            # Charger le fichier Excel
            wb = openpyxl.load_workbook(excel_path, data_only=True)
            
            # 1. Valider les onglets (Sxx)
            sheet_pattern = re.compile(self.schema['sheet_pattern'])
            for sheet_name in wb.sheetnames:
                if sheet_pattern.match(sheet_name):
                    result['sheets_found'].append(sheet_name)
            
            if not result['sheets_found']:
                result['errors'].append("❌ Aucun onglet de semaine trouvé (format Sxx attendu)")
                result['valid'] = False
                return result
            
            # 2. Valider une feuille (S48 ou la plus récente)
            if 'S48' in result['sheets_found']:
                ws = wb['S48']
                week_to_validate = 'S48'
            else:
                # Prendre la semaine la plus récente
                weeks = sorted(result['sheets_found'], key=lambda x: int(x[1:]))
                week_to_validate = weeks[-1]
                ws = wb[week_to_validate]
            
            logger.info(f"🔍 Validation de l'onglet : {week_to_validate}")
            
            # 3. Valider les en-têtes (ligne 3)
            header_row = self.schema['header_row']
            
            for col_letter, col_schema in self.schema['columns'].items():
                col_idx = openpyxl.utils.column_index_from_string(col_letter)
                cell_value = ws.cell(row=header_row, column=col_idx).value
                
                expected_name = col_schema['expected_name']
                alternate_names = col_schema.get('alternate_names', [])
                is_ignored = col_schema.get('ignored', False)
                is_required = col_schema.get('required', False)
                
                # Comparer avec le nom attendu et les noms alternatifs
                match = self._compare_column_name(cell_value, expected_name, alternate_names)
                is_alternate = False
                matched_alternate = None
                
                # Vérifier si c'est un nom alternatif (si pas de match direct)
                if alternate_names and not match:
                    for alt_name in alternate_names:
                        if self._compare_column_name(cell_value, alt_name, []):
                            match = True
                            is_alternate = True
                            matched_alternate = alt_name
                            break
                
                if match:
                    # Colonne trouvée (nom exact ou alternatif accepté)
                    if is_alternate:
                        # Nom alternatif détecté et accepté
                        result['warnings'].append(
                            f"ℹ️ Col {col_letter}: Nom alternatif accepté '{cell_value}' (attendu '{expected_name}')"
                        )
                        
                    result['columns_validated'][col_letter] = {
                        "status": "✅" if not is_alternate else "✅ (nom alternatif)",
                        "found": cell_value,
                        "matched_alternate": matched_alternate,
                        "expected": expected_name,
                        "db_field": col_schema.get('db_field')
                    }
                else:
                    # Colonne ne correspond pas
                    if is_ignored:
                        # Colonne ignorée, juste un warning
                        result['warnings'].append(
                            f"⚠️ Col {col_letter}: Attendu '{expected_name}', trouvé '{cell_value}' (colonne ignorée)"
                        )
                        result['columns_validated'][col_letter] = {
                            "status": "⚠️",
                            "found": cell_value,
                            "expected": expected_name,
                            "db_field": None
                        }
                    elif is_required:
                        # Colonne requise manquante ou différente
                        result['errors'].append(
                            f"❌ Col {col_letter} (REQUISE): Attendu '{expected_name}', trouvé '{cell_value}'"
                        )
                        result['valid'] = False
                        result['columns_validated'][col_letter] = {
                            "status": "❌",
                            "found": cell_value,
                            "expected": expected_name,
                            "db_field": col_schema.get('db_field')
                        }
                    else:
                        # Colonne optionnelle différente
                        result['warnings'].append(
                            f"⚠️ Col {col_letter}: Attendu '{expected_name}', trouvé '{cell_value}'"
                        )
                        result['columns_validated'][col_letter] = {
                            "status": "⚠️",
                            "found": cell_value,
                            "expected": expected_name,
                            "db_field": col_schema.get('db_field')
                        }
            
            # 4. Résumé
            logger.info(f"✅ Validation terminée : {len(result['errors'])} erreurs, {len(result['warnings'])} warnings")
            
        except FileNotFoundError:
            result['errors'].append(f"❌ Fichier Excel introuvable : {excel_path}")
            result['valid'] = False
        except Exception as e:
            result['errors'].append(f"❌ Erreur lors de la validation : {str(e)}")
            result['valid'] = False
            logger.exception(e)
        
        return result
    
    def get_column_mapping(self) -> Dict[str, str]:
        """
        Retourne le mapping complet colonne Excel → champ base de données.
        
        Returns:
            Dict {col_letter: db_field}
        """
        mapping = {}
        for col_letter, col_schema in self.schema['columns'].items():
            db_field = col_schema.get('db_field')
            if db_field and not col_schema.get('ignored', False):
                mapping[col_letter] = db_field
        return mapping
    
    def get_required_columns(self) -> List[str]:
        """
        Retourne la liste des colonnes obligatoires.
        
        Returns:
            Liste de lettres de colonnes (ex: ['A', 'B', 'C', 'D'])
        """
        required = []
        for col_letter, col_schema in self.schema['columns'].items():
            if col_schema.get('required', False):
                required.append(col_letter)
        return required

